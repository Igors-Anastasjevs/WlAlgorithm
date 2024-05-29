import openpyxl as xl
import time
def generategraphs(graphgenerator, range):
    '''
    Generates desired type of graphs
    :param graphgenerator: function which creates graph, using only 1 parameter.
    :param range: class range, which consists of numbers of desired nodes in graphs.
    :return: dictionary with number of nodes as a key and graph as a value.
    '''
    graphs = dict()
    for i in range:
        graph = graphgenerator(i)
        graphs.update({i:graph})
    return graphs
def savedata(filename, graphs:dict, graphisomorthismfunction):
    '''
    This is a function, which at first measures performance of an input function 20 times, and then saves collected data
    into .xlsx file.
    :param filename: string, which has a .xlsx-file name, where data will be recorded into.
    :param graphs: dictionary, which has integer as number of nodes in a graph as a key, and graphs themselves as a value
    :param graphisomorthismfunction: a function pointer for desired function to be used in order to measure performance.
    '''
    data = dict()
    for size in graphs.keys():
        G = graphs[size]
        H = G.copy()
        data.update({size: []})
        for i in range(20):
            start = time.time()
            graphisomorthismfunction(G, H)
            end = time.time()
            data[size].append(end - start)
        print('finished ', size)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'data'
    i = 1
    for key in data.keys():
        ws.cell(row=i, column=1, value=key)
        for j in range(20):
            ws.cell(row=i, column=j + 2, value=data[key][j])
        i += 1
    wb.save(filename=filename)


